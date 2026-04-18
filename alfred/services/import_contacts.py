"""Bulk contact import via CSV or XLSX."""
import csv
import io
from typing import Any

import httpx
import structlog

from alfred.config import settings
from alfred.db.client import get_db
from alfred.services.contacts import create_contact_confirmed, find_similar_contacts

log = structlog.get_logger()

VALID_COLUMNS = {"display_name", "company", "role", "cadence_days", "relationship_type", "tags", "how_we_met"}
REQUIRED_COLUMNS = {"display_name"}
VALID_RELATIONSHIP_TYPES = {"friend", "professional", "family", "other"}
MAX_ROWS = 100


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def build_template_csv() -> bytes:
    """Generate a template CSV with one example row."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["display_name", "company", "role", "cadence_days", "relationship_type", "tags", "how_we_met"])
    writer.writerow(["João Silva", "Empresa X", "CEO", "15", "professional", "cliente|vip", "Evento Tech 2024"])
    writer.writerow(["Maria Santos", "Empresa Y", "Diretora", "30", "friend", "amiga", ""])
    # utf-8-sig adds BOM so Excel opens correctly without encoding issues
    return output.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# Shared validation helpers
# ---------------------------------------------------------------------------

def _validate_columns(fieldnames: set[str]) -> list[str]:
    """Validate column names; return list of error strings (empty = ok)."""
    errors: list[str] = []

    unknown = fieldnames - VALID_COLUMNS
    if unknown:
        errors.append(
            f"Coluna(s) inválida(s): {', '.join(sorted(unknown))}. "
            f"Válidas: {', '.join(sorted(VALID_COLUMNS))}."
        )

    missing = REQUIRED_COLUMNS - fieldnames
    if missing:
        errors.append(f"Coluna obrigatória ausente: {', '.join(sorted(missing))}.")

    return errors


def _validate_rows(raw_rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[str]]:
    """Validate and coerce a list of raw string-valued dicts into contact dicts.

    Returns (rows, errors). Uses ``int(float(...))`` for cadence_days so that
    Excel numeric cells (which arrive as floats) are handled correctly.
    """
    errors: list[str] = []
    rows: list[dict[str, Any]] = []

    if not raw_rows:
        return [], ["Arquivo não tem dados além do cabeçalho."]

    if len(raw_rows) > MAX_ROWS:
        return [], [
            f"Máximo de {MAX_ROWS} contatos por importação. "
            f"Seu arquivo tem {len(raw_rows)} linhas."
        ]

    for i, raw_row in enumerate(raw_rows, start=2):  # line/row 1 = header
        display_name = raw_row.get("display_name", "")
        if not display_name:
            errors.append(f"Linha {i}: display_name é obrigatório e não pode estar vazio.")
            continue

        # Validate cadence_days
        cadence_days = 15
        cadence_raw = raw_row.get("cadence_days", "")
        if cadence_raw:
            try:
                # int(float(...)) handles cells that Excel stores as floats (e.g. "30.0")
                cadence_days = int(float(cadence_raw))
                if not (1 <= cadence_days <= 365):
                    errors.append(
                        f"Linha {i} ({display_name}): cadence_days deve ser entre 1 e 365, "
                        f"recebeu '{cadence_raw}'."
                    )
                    continue
            except ValueError:
                errors.append(
                    f"Linha {i} ({display_name}): cadence_days deve ser um número inteiro, "
                    f"recebeu '{cadence_raw}'."
                )
                continue

        # Validate relationship_type
        relationship_type: str | None = raw_row.get("relationship_type") or None
        if relationship_type and relationship_type not in VALID_RELATIONSHIP_TYPES:
            errors.append(
                f"Linha {i} ({display_name}): relationship_type inválido '{relationship_type}'. "
                f"Válidos: {', '.join(sorted(VALID_RELATIONSHIP_TYPES))}."
            )
            continue

        # Parse pipe-separated tags
        tags_raw = raw_row.get("tags", "")
        tags = [t.strip() for t in tags_raw.split("|") if t.strip()] if tags_raw else []

        contact: dict[str, Any] = {"display_name": display_name, "cadence_days": cadence_days}
        if raw_row.get("company"):
            contact["company"] = raw_row["company"]
        if raw_row.get("role"):
            contact["role"] = raw_row["role"]
        if relationship_type:
            contact["relationship_type"] = relationship_type
        if tags:
            contact["tags"] = tags
        if raw_row.get("how_we_met"):
            contact["how_we_met"] = raw_row["how_we_met"]

        rows.append(contact)

    return rows, errors


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def parse_and_validate(csv_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse and validate CSV bytes.

    Returns (rows, errors). If errors is non-empty, do not import.
    """
    # Attempt decoding (handle BOM and latin-1 fallback)
    try:
        text = csv_bytes.decode("utf-8-sig").strip()
    except UnicodeDecodeError:
        try:
            text = csv_bytes.decode("latin-1").strip()
        except UnicodeDecodeError:
            return [], ["Não foi possível ler o arquivo. Certifique-se que está salvo em UTF-8."]

    if not text:
        return [], ["Arquivo CSV vazio."]

    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        return [], ["CSV não tem cabeçalho."]

    fieldnames_lower = {f.strip().lower() for f in reader.fieldnames}
    col_errors = _validate_columns(fieldnames_lower)
    if col_errors:
        return [], col_errors

    raw_rows = list(reader)
    if not raw_rows:
        return [], ["CSV não tem dados além do cabeçalho."]

    # Normalise keys and values to lowercase stripped strings
    normalised: list[dict[str, str]] = [
        {k.strip().lower(): (v.strip() if v else "") for k, v in row.items() if k}
        for row in raw_rows
    ]

    return _validate_rows(normalised)


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

def parse_xlsx(xlsx_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse and validate XLSX bytes.

    Returns (rows, errors). If errors is non-empty, do not import.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ["openpyxl não está instalado. Execute: pip install openpyxl"]

    try:
        wb = load_workbook(filename=io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"Não foi possível abrir o arquivo Excel: {exc}"]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Arquivo Excel vazio."]

    fieldnames = [str(h).strip().lower() if h is not None else "" for h in header_row]
    fieldnames_set = set(fieldnames) - {""}

    col_errors = _validate_columns(fieldnames_set)
    if col_errors:
        return [], col_errors

    # Build normalised dicts from remaining rows
    normalised: list[dict[str, str]] = []
    for row in rows_iter:
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue
        row_dict: dict[str, str] = {}
        for col_name, cell_val in zip(fieldnames, row):
            if not col_name:
                continue
            if cell_val is None:
                row_dict[col_name] = ""
            else:
                row_dict[col_name] = str(cell_val).strip()
        normalised.append(row_dict)

    wb.close()

    return _validate_rows(normalised)


# ---------------------------------------------------------------------------
# Preview helper
# ---------------------------------------------------------------------------

def build_preview(rows: list[dict[str, Any]]) -> str:
    n = len(rows)
    s = "s" if n != 1 else ""
    lines = [f"📋 *Arquivo válido — {n} contato{s} encontrado{s}:*\n"]

    for c in rows[:10]:
        parts = []
        if c.get("company"):
            parts.append(c["company"])
        if c.get("role"):
            parts.append(c["role"])
        if c.get("relationship_type"):
            parts.append(c["relationship_type"])
        parts.append(f"{c.get('cadence_days', 15)}d")
        lines.append(f"• {c['display_name']} — {' · '.join(parts)}")

    if n > 10:
        lines.append(f"_...e mais {n - 10} contatos_")

    lines.append("\nConfirma a importação?")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Download helper
# ---------------------------------------------------------------------------

async def download_file(file_id: str) -> bytes:
    """Download a file from Telegram by file_id."""
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"https://api.telegram.org/bot{settings.telegram_bot_token}/getFile",
            params={"file_id": file_id},
        )
        resp.raise_for_status()
        file_path = resp.json()["result"]["file_path"]

        file_resp = await client.get(
            f"https://api.telegram.org/file/bot{settings.telegram_bot_token}/{file_path}",
        )
        file_resp.raise_for_status()
        return file_resp.content


# Backwards-compat alias
download_csv = download_file


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------

async def bulk_import(user_id: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Create contacts, skipping duplicates (via find_similar_contacts).

    Returns {"created": N, "skipped": [names]}.
    """
    created = 0
    skipped: list[str] = []

    for contact in rows:
        name = contact["display_name"]
        similar = await find_similar_contacts(user_id=user_id, display_name=name)
        if similar:
            log.info("import.skipped_duplicate", user_id=user_id, display_name=name)
            skipped.append(name)
            continue
        await create_contact_confirmed(user_id=user_id, **contact)
        created += 1

    log.info("import.completed", user_id=user_id, created=created, skipped=len(skipped))
    return {"created": created, "skipped": skipped}


# ---------------------------------------------------------------------------
# Duplicate detection (Task 3)
# ---------------------------------------------------------------------------

async def check_duplicates(
    user_id: str,
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Separate rows into clean (no match) and duplicates (with existing match).

    Each duplicate entry: ``{"csv_row": {...}, "existing": {"id", "display_name", "company"}}``.
    Uses ``find_similar_contacts`` from ``alfred.services.contacts``.
    """
    clean: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []

    for row in rows:
        name = row["display_name"]
        similar = await find_similar_contacts(user_id=user_id, display_name=name)
        if similar:
            duplicates.append({"csv_row": row, "existing": similar[0]})
        else:
            clean.append(row)

    log.info(
        "import.check_duplicates",
        user_id=user_id,
        clean=len(clean),
        duplicates=len(duplicates),
    )
    return clean, duplicates


def build_grouped_preview(
    clean: list[dict[str, Any]],
    duplicates: list[dict[str, Any]],
) -> str:
    """Build a grouped preview: ✅ N novos + ⚠️ M duplicatas."""
    lines: list[str] = []

    n_clean = len(clean)
    if n_clean:
        s = "s" if n_clean != 1 else ""
        lines.append(f"✅ *{n_clean} contato{s} novo{s}:*")
        for c in clean[:10]:
            parts: list[str] = []
            if c.get("company"):
                parts.append(c["company"])
            if c.get("role"):
                parts.append(c["role"])
            suffix = f" — {' · '.join(parts)}" if parts else ""
            lines.append(f"• {c['display_name']}{suffix}")
        if n_clean > 10:
            lines.append(f"_...e mais {n_clean - 10}_")

    n_dup = len(duplicates)
    if n_dup:
        if lines:
            lines.append("")
        s = "s" if n_dup != 1 else ""
        lines.append(f"⚠️ *{n_dup} duplicata{s} detectada{s}:*")
        for dup in duplicates:
            csv_name = dup["csv_row"]["display_name"]
            existing = dup["existing"]
            ex_name = existing.get("display_name", "?")
            ex_company = existing.get("company") or ""
            ex_label = f"{ex_name}" + (f" ({ex_company})" if ex_company else "")
            lines.append(f"• {csv_name} → existente: {ex_label}")

    return "\n".join(lines)


def build_duplicate_comparison(dup: dict[str, Any], index: int, total: int) -> str:
    """Side-by-side comparison for a single duplicate: CSV data vs existing data."""
    csv_row = dup["csv_row"]
    existing = dup["existing"]

    csv_name = csv_row.get("display_name", "?")
    ex_name = existing.get("display_name", "?")

    lines = [f"📊 *Duplicata {index}/{total} — {csv_name}*\n"]

    lines.append("📄 *CSV (novo):*")
    for field in ("display_name", "company", "role", "relationship_type", "how_we_met", "cadence_days", "tags"):
        val = csv_row.get(field)
        if val:
            lines.append(f"  • {field}: {val}")

    lines.append("\n👤 *Existente no Alfred:*")
    lines.append(f"  • display\\_name: {ex_name}")
    if existing.get("company"):
        lines.append(f"  • company: {existing['company']}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Merge, replace, execute_import (Task 4)
# ---------------------------------------------------------------------------

MERGEABLE_FIELDS = ["company", "role", "relationship_type", "how_we_met", "tags"]


async def merge_contact(user_id: str, existing_id: str, csv_row: dict[str, Any]) -> str:
    """Fetch existing contact and update only empty fields with CSV data."""
    db = get_db()
    result = (
        db.table("contacts")
        .select("id, display_name, company, role, relationship_type, how_we_met, tags")
        .eq("user_id", user_id)
        .eq("id", existing_id)
        .single()
        .execute()
    )
    existing = result.data

    updates: dict[str, Any] = {}
    for field in MERGEABLE_FIELDS:
        existing_val = existing.get(field)
        csv_val = csv_row.get(field)
        # Only fill if existing is empty/None and CSV has a value
        if not existing_val and csv_val:
            updates[field] = csv_val

    if updates:
        db.table("contacts").update(updates).eq("user_id", user_id).eq("id", existing_id).execute()
        log.info("import.merge_contact", user_id=user_id, contact_id=existing_id, fields=list(updates.keys()))
    else:
        log.info("import.merge_contact.no_changes", user_id=user_id, contact_id=existing_id)

    name = existing.get("display_name", existing_id)
    return f"Contato **{name}** mesclado (campos preenchidos: {', '.join(updates.keys()) or 'nenhum'})."


async def replace_contact(user_id: str, existing_id: str, csv_row: dict[str, Any]) -> str:
    """Overwrite existing contact fields with all CSV data."""
    updates: dict[str, Any] = {}
    for field in MERGEABLE_FIELDS:
        csv_val = csv_row.get(field)
        if csv_val is not None:
            updates[field] = csv_val

    # Also update display_name and cadence_days if present
    for field in ("display_name", "cadence_days"):
        if csv_row.get(field):
            updates[field] = csv_row[field]

    db = get_db()
    db.table("contacts").update(updates).eq("user_id", user_id).eq("id", existing_id).execute()
    log.info("import.replace_contact", user_id=user_id, contact_id=existing_id)

    name = csv_row.get("display_name", existing_id)
    return f"Contato **{name}** substituído com dados do CSV."


async def execute_import(
    user_id: str,
    clean_rows: list[dict[str, Any]],
    duplicates: list[dict[str, Any]],
    decisions: dict[str, Any],
) -> dict[str, Any]:
    """Execute full import.

    Decisions are read from ``dup["decision"]`` (embedded in each duplicate dict).
    Valid values: "skip", "import_new", "merge", "replace".

    Returns {"created", "skipped", "merged", "replaced", "merged_details"}.
    """
    created = 0
    skipped = 0
    merged = 0
    replaced = 0
    merged_details: list[str] = []

    # Import all clean rows
    for row in clean_rows:
        await create_contact_confirmed(user_id=user_id, **row)
        created += 1

    # Apply decisions for duplicates
    for dup in duplicates:
        decision = dup.get("decision") or decisions.get(dup["csv_row"]["display_name"], "skip")
        csv_row = dup["csv_row"]
        existing_id = dup["existing"]["id"]

        if decision == "skip":
            skipped += 1
        elif decision == "import_new":
            await create_contact_confirmed(user_id=user_id, **csv_row)
            created += 1
        elif decision == "merge":
            result_msg = await merge_contact(user_id, existing_id, csv_row)
            merged += 1
            merged_details.append(result_msg)
        elif decision == "replace":
            await replace_contact(user_id, existing_id, csv_row)
            replaced += 1
        else:
            # Default: skip unknown decisions
            skipped += 1

    log.info(
        "import.execute_import",
        user_id=user_id,
        created=created,
        skipped=skipped,
        merged=merged,
        replaced=replaced,
    )
    return {
        "created": created,
        "skipped": skipped,
        "merged": merged,
        "replaced": replaced,
        "merged_details": merged_details,
    }


def build_import_report(result: dict[str, Any]) -> str:
    """Build a final import report in MarkdownV2 format."""
    created = result.get("created", 0)
    skipped = result.get("skipped", 0)
    merged = result.get("merged", 0)
    replaced = result.get("replaced", 0)

    lines = [r"✅ *Importação concluída\!*", ""]

    if created:
        s = "s" if created != 1 else ""
        lines.append(f"• {created} contato{s} criado{s}")
    if merged:
        s = "s" if merged != 1 else ""
        lines.append(f"• {merged} contato{s} mesclado{s}")
    if replaced:
        s = "s" if replaced != 1 else ""
        lines.append(f"• {replaced} contato{s} substituído{s}")
    if skipped:
        s = "s" if skipped != 1 else ""
        lines.append(f"• {skipped} duplicata{s} ignorada{s}")

    return "\n".join(lines)
